from aiogram import Router, types, F
from aiogram.fsm.context import FSMContext
from app.database.models import User, UserRole
from app.container import Container

from app.services.product_service import ProductService
from app.services.user_service import UserService
from app.services.category_service import CategoryService
from app.services.transaction_service import TransactionService

from app.telegram.states.admin import AddProductState, AddCategoryState, WaitAdminReply, EditStaffProfileState
from app.telegram.keyboards.admin import (
    main_admin_kb, 
    products_list_kb, 
    cancel_kb,
    cancel_admin_inline_kb,
    product_edit_kb, 
    categories_list_kb, 
    approve_user_kb,
    stats_periods_kb,
    undo_tx_kb
)

router = Router()

# Filter for admin role only
# We can do this with a Custom Filter, but for simplicity we'll check inside or use F.role == UserRole.ADMIN
# Since db_user is passed to every handler, we can filter using it (Aiogram 3 allows checking kwargs in F-magic)
# However, the easiest way for now is a custom function or simple lambda.
router.message.filter(lambda event, db_user=None: db_user is not None and db_user.role == UserRole.ADMIN)
router.callback_query.filter(lambda event, db_user=None: db_user is not None and db_user.role == UserRole.ADMIN)

@router.message(F.text == "📦 Ombor")
async def show_stock(message: types.Message, container: Container):
    product_service: ProductService = container.get("product_service")
    products = await product_service.get_all_products()
    
    if not products:
        await message.answer("Ombor bo'sh.", reply_markup=products_list_kb(products))
        return
        
    await message.answer("📦 Ombordagi mahsulotlar ro'yxati:", reply_markup=products_list_kb(products))

@router.callback_query(F.data == "back_to_stock")
async def cb_back_to_stock(call: types.CallbackQuery, container: Container):
    product_service: ProductService = container.get("product_service")
    products = await product_service.get_all_products()
    await call.message.edit_text("📦 Ombordagi mahsulotlar ro'yxati:", reply_markup=products_list_kb(products))

@router.message(F.text == "Bekor qilish")
async def cancel_handler(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Amal bekor qilindi.", reply_markup=main_admin_kb())

@router.callback_query(F.data == "admin_cancel")
async def admin_cancel_cb(call: types.CallbackQuery, state: FSMContext):
    """Inline-отмена для админа — редактирует сообщение, не трогает нижнее меню."""
    await state.clear()
    try:
        await call.message.edit_text("❌ Amal bekor qilindi.")
    except Exception:
        await call.message.delete()
        await call.message.answer("❌ Amal bekor qilindi.", reply_markup=main_admin_kb())

# --- Category Management ---
@router.message(F.text == "🗂 Kategoriyalar")
async def show_categories(message: types.Message, container: Container):
    from app.services.category_service import CategoryService
    from app.telegram.keyboards.admin import categories_list_kb
    category_service: CategoryService = container.get("category_service")
    categories = await category_service.get_all_categories()
    
    if not categories:
        await message.answer("Kategoriyalar hozircha yo'q. Ularni qo'shishingiz mumkin.", reply_markup=categories_list_kb([]))
    else:
        await message.answer("Kategoriyalarni boshqarish:", reply_markup=categories_list_kb(categories))

@router.callback_query(F.data == "add_category")
async def cb_add_category(call: types.CallbackQuery, state: FSMContext):
    from app.telegram.states.admin import AddCategoryState
    await state.set_state(AddCategoryState.name)
    await call.message.edit_text("Yangi kategoriya nomini kiriting:", reply_markup=cancel_admin_inline_kb())

@router.message(AddCategoryState.name)
async def process_add_category_name(message: types.Message, state: FSMContext, container: Container):
    category_name = message.text.strip()
    category_service: CategoryService = container.get("category_service")
    
    try:
        await category_service.create_category(category_name)
        await message.answer(f"✅ '{category_name}' kategoriyasi muvaffaqiyatli yaratildi!", reply_markup=main_admin_kb())
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}", reply_markup=main_admin_kb())
    finally:
        await state.clear()

@router.callback_query(F.data.startswith("manage_cat_"))
async def cb_manage_category(call: types.CallbackQuery):
    cat_id = int(call.data.split("_")[2])
    await call.answer(f"Kategoriya ID {cat_id} ni boshqarish (ishlanmoqda)", show_alert=True)

# --- Add Product ---
@router.callback_query(F.data == "add_product")
async def start_add_product(call: types.CallbackQuery, state: FSMContext, container: Container):
    from app.services.category_service import CategoryService
    from app.telegram.keyboards.admin import categories_list_kb
    category_service: CategoryService = container.get("category_service")
    categories = await category_service.get_all_categories()
    
    if not categories:
        await call.answer("Avvalo kategoriya yarating (🗂 Kategoriyalar)!", show_alert=True)
        return
        
    await state.set_state(AddProductState.category_id)
    await call.message.edit_text("Yangi mahsulot uchun kategoriya tanlang:", reply_markup=categories_list_kb(categories, for_selection=True))

@router.callback_query(AddProductState.category_id, F.data.startswith("select_cat_"))
async def cb_select_category_for_product(call: types.CallbackQuery, state: FSMContext):
    cat_id = int(call.data.split("_")[2])
    await state.update_data(category_id=cat_id)
    await state.set_state(AddProductState.name)
    await call.message.edit_text("Ajoyib. Yangi mahsulot nomini kiriting:", reply_markup=cancel_admin_inline_kb())

@router.message(AddProductState.name)
async def process_product_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("Mahsulot narxini kiriting (raqam):")
    await state.set_state(AddProductState.price)

@router.message(AddProductState.price)
async def process_product_price(message: types.Message, state: FSMContext):
    try:
        price = float(message.text.replace(",", "."))
    except ValueError:
        await message.answer("Iltimos, narx uchun to'g'ri raqam kiriting.")
        return
        
    await state.update_data(price=price)
    await message.answer("Ombordagi boshlang'ich miqdorni kiriting:")
    await state.set_state(AddProductState.initial_quantity)

@router.message(AddProductState.initial_quantity)
async def process_product_quantity(message: types.Message, state: FSMContext, container: Container):
    try:
        qty = int(message.text)
    except ValueError:
        await message.answer("Iltimos, butun son kiriting.")
        return
        
    data = await state.get_data()
    product_service: ProductService = container.get("product_service")
    
    try:
        cat_id = data.get("category_id")
        await product_service.create_product(name=data['name'], price=data['price'], quantity=qty, category_id=cat_id)
        await message.answer(f"✅ '{data['name']}' mahsuloti qo'shildi!", reply_markup=main_admin_kb())
    except Exception as e:
        await message.answer("❌ MBga qo'shishda xatolik. Ehtimol bunday nomdagi mahsulot mavjud.", reply_markup=main_admin_kb())
        
    await state.clear()


# --- Edit Product Stock ---
@router.callback_query(F.data.startswith("prod_edit_"))
async def cb_edit_product(call: types.CallbackQuery, container: Container):
    product_id = int(call.data.split("_")[2])
    product_service: ProductService = container.get("product_service")
    
    product = await product_service.get_product_by_id(product_id)
    if product:
        text = f"📦 <b>Mahsulot:</b> {product.name}\nNarx: {product.price:,} so'm\nQoldiq: <b>{product.quantity}</b> dona"
        await call.message.edit_text(text, parse_mode="HTML", reply_markup=product_edit_kb(product_id))

@router.callback_query(F.data.startswith("prod_inc_"))
async def cb_inc_product(call: types.CallbackQuery, container: Container):
    product_id = int(call.data.split("_")[2])
    product_service: ProductService = container.get("product_service")
    
    product = await product_service.update_quantity(product_id, 1)
    if product:
        text = f"📦 <b>Mahsulot:</b> {product.name}\nNarx: {product.price:,} so'm\nQoldiq: <b>{product.quantity}</b> dona"
        await call.message.edit_text(text, parse_mode="HTML", reply_markup=product_edit_kb(product_id))
    else:
        await call.answer("Mahsulot topilmadi", show_alert=True)

@router.callback_query(F.data.startswith("prod_dec_"))
async def cb_dec_product(call: types.CallbackQuery, container: Container):
    product_id = int(call.data.split("_")[2])
    product_service: ProductService = container.get("product_service")
    
    # Check if quantity > 0
    product = await product_service.get_product_by_id(product_id)
    if product and product.quantity > 0:
        new_product = await product_service.update_quantity(product_id, -1)
        text = f"📦 <b>Mahsulot:</b> {new_product.name}\nNarx: {new_product.price:,} so'm\nQoldiq: <b>{new_product.quantity}</b> dona"
        await call.message.edit_text(text, parse_mode="HTML", reply_markup=product_edit_kb(product_id))
    else:
        await call.answer("Kamaytirib bo'lmaydi (qoldiq 0 yoki mahsulot topilmadi)", show_alert=True)

@router.callback_query(F.data.startswith("prod_del_conf_"))
async def cb_delete_product_conf(call: types.CallbackQuery, container: Container):
    product_id = int(call.data.split("_")[3])
    from app.telegram.keyboards.admin import product_delete_confirm_kb
    product_service: ProductService = container.get("product_service")
    product = await product_service.get_product_by_id(product_id)
    if product:
        await call.message.edit_text(f"⚠️ <b>{product.name}</b> mahsulotini o'chirishga ishonchingiz komilmi?\nBu amalni ortga qaytarib bo'lmaydi.", parse_mode="HTML", reply_markup=product_delete_confirm_kb(product_id))

@router.callback_query(F.data.startswith("prod_del_yes_"))
async def cb_delete_product_yes(call: types.CallbackQuery, container: Container):
    product_id = int(call.data.split("_")[3])
    product_service: ProductService = container.get("product_service")
    
    success = await product_service.delete_product(product_id)
    if success:
        await call.answer("Mahsulot o'chirildi", show_alert=True)
        products = await product_service.get_all_products()
        await call.message.edit_text("📦 Ombordagi mahsulotlar ro'yxati:", reply_markup=products_list_kb(products))
    else:
        await call.answer("O'chirishda xatolik", show_alert=True)

# Update start command for admin
@router.message(F.text == "/start")
async def admin_start(message: types.Message, db_user: User):
    await message.answer(f"Administrator paneliga xush kelibsiz, {db_user.username}!", reply_markup=main_admin_kb())

# --- Staff Moderation ---
@router.callback_query(F.data.startswith("approve_"))
async def cb_approve_user(call: types.CallbackQuery, container: Container):
    user_id = int(call.data.split("_")[1])
    user_service: UserService = container.get("user_service")
    
    updated_user = await user_service.update_user_role(user_id, UserRole.WORKER)
    if updated_user:
        await call.message.edit_text(f"✅ @{updated_user.username or user_id} foydalanuvchisi ruxsat oldi (WORKER).")
        try:
            await call.bot.send_message(user_id, "🎉 Sizning so'rovingiz tasdiqlandi! Endi sizga xodim menyusi ruxsat etildi.\n/start tugmasini bosing.")
        except Exception:
            pass
    else:
        await call.answer("Rolni yangilashda xatolik", show_alert=True)

@router.callback_query(F.data.startswith("reject_"))
async def cb_reject_user(call: types.CallbackQuery, container: Container):
    user_id = int(call.data.split("_")[1])
    user_service: UserService = container.get("user_service")
    
    updated_user = await user_service.update_user_role(user_id, UserRole.BANNED)
    if updated_user:
        await call.message.edit_text(f"⛔ @{updated_user.username or user_id} foydalanuvchi rad etildi (BANNED).")
    else:
        await call.answer("Rolni yangilashda xatolik", show_alert=True)

# --- Manage Staff ---
@router.message(F.text == "👥 Xodimlar")
async def show_staff(message: types.Message, container: Container):
    user_service: UserService = container.get("user_service")
    
    users = await user_service.get_all_users()
    # List Workers, Pending, and BANNED to make management easier
    workers = [u for u in users if u.role in [UserRole.WORKER, UserRole.PENDING, UserRole.BANNED]]
    
    if not workers:
        text = "Xodimlar va so'rovlar hozircha yo'q."
        await message.answer(text, reply_markup=main_admin_kb())
    else:
        from app.telegram.keyboards.admin import staff_list_kb
        text = "👥 Boshqarish uchun xodim yoki nomzodni tanlang:"
        await message.answer(text, reply_markup=staff_list_kb(workers))

@router.callback_query(F.data == "staff_list")
async def cb_staff_list(call: types.CallbackQuery, container: Container):
    user_service: UserService = container.get("user_service")
    users = await user_service.get_all_users()
    workers = [u for u in users if u.role in [UserRole.WORKER, UserRole.PENDING, UserRole.BANNED]]
    
    if not workers:
        await call.message.edit_text("Xodimlar va so'rovlar hozircha yo'q.", reply_markup=None)
        return
        
    from app.telegram.keyboards.admin import staff_list_kb
    text = "👥 Boshqarish uchun xodim yoki nomzodni tanlang:"
    await call.message.edit_text(text, reply_markup=staff_list_kb(workers))

@router.callback_query(F.data.startswith("staff_profile_"))
async def cb_staff_profile(call: types.CallbackQuery, container: Container):
    tg_id = int(call.data.split("_")[2])
    user_service: UserService = container.get("user_service")
    user = await user_service.get_user_by_tg_id(tg_id)
    if not user:
        await call.answer("Foydalanuvchi topilmadi.", show_alert=True)
        return
        
    from app.telegram.keyboards.admin import staff_profile_kb
    
    from app.telegram.keyboards.worker import kpi_progress_bar
    import html
    
    # Calculate current progress for KPI
    transaction_service: TransactionService = container.get("transaction_service")
    stats_today = await transaction_service.get_admin_statistics("today", user_id=user.id)
    revenue_today = sum(t.total_price for t in stats_today)
    
    status = "✅ Faol" if user.is_active else "⛔ Bloklangan"
    if user.role == UserRole.PENDING:
        status = "⏳ Tasdiqlash kutilmoqda"
        
    kpi_bar = kpi_progress_bar(revenue_today, user.kpi) if user.kpi > 0 else "🎯 KPI belgilanmagan"
    
    username = html.escape(user.username) if user.username else "---"
    full_name = html.escape(user.full_name) if user.full_name else "---"
    phone = html.escape(user.phone) if user.phone else "---"
    
    text = (
        f"👤 <b>Profil:</b> {full_name}\n\n"
        f"<blockquote>• ID: <code>{user.tg_id}</code>\n"
        f"• Username: @{username}\n"
        f"• Telefon: {phone}\n"
        f"• Rol: {user.role.value}\n"
        f"• Holat: {status}\n"
        f"• Ro'yxatdan o'tgan: {user.joined_at.strftime('%Y-%m-%d') if user.joined_at else '---'}</blockquote>\n\n"
        f"🎯 <b>Bugungi KPI:</b>\n"
        f"• Maqsad: {user.kpi:,} so'm\n"
        f"• Daromad: {round(revenue_today, 2):,} so'm\n\n"
        f"{kpi_bar}\n\n"
        f"Amalni tanlang:"
    )
    
    await call.message.edit_text(text, parse_mode="HTML", reply_markup=staff_profile_kb(tg_id, user.role))

@router.callback_query(F.data.startswith("staff_edit_name_"))
async def cb_staff_edit_name(call: types.CallbackQuery, state: FSMContext):
    tg_id = int(call.data.split("_")[3])
    await state.update_data(target_tg_id=tg_id)
    await state.set_state(EditStaffProfileState.full_name)
    await call.message.edit_text("Xodim F.I.Sh ni kiriting:", reply_markup=cancel_admin_inline_kb())

@router.message(EditStaffProfileState.full_name)
async def process_edit_staff_name(message: types.Message, state: FSMContext, container: Container):
    data = await state.get_data()
    tg_id = data.get("target_tg_id")
    user_service: UserService = container.get("user_service")
    await user_service.update_user_profile(tg_id, full_name=message.text.strip())
    await message.answer(f"✅ F.I.Sh yangilandi!", reply_markup=main_admin_kb())
    await state.clear()

@router.callback_query(F.data.startswith("staff_edit_phone_"))
async def cb_staff_edit_phone(call: types.CallbackQuery, state: FSMContext):
    tg_id = int(call.data.split("_")[3])
    await state.update_data(target_tg_id=tg_id)
    await state.set_state(EditStaffProfileState.phone)
    await call.message.edit_text("Xodim telefon raqamini kiriting:", reply_markup=cancel_admin_inline_kb())

@router.message(EditStaffProfileState.phone)
async def process_edit_staff_phone(message: types.Message, state: FSMContext, container: Container):
    data = await state.get_data()
    tg_id = data.get("target_tg_id")
    user_service: UserService = container.get("user_service")
    await user_service.update_user_profile(tg_id, phone=message.text.strip())
    await message.answer(f"✅ Telefon yangilandi!", reply_markup=main_admin_kb())
    await state.clear()

from app.telegram.states.admin import EditStaffKPIState

@router.callback_query(F.data.startswith("staff_edit_kpi_"))
async def cb_staff_edit_kpi(call: types.CallbackQuery, state: FSMContext):
    tg_id = int(call.data.split("_")[3])
    await state.update_data(target_tg_id=tg_id)
    await state.set_state(EditStaffKPIState.kpi)
    await call.message.edit_text("Yangi KPI kiriting (so'm yoki donada maqsadli qiymat):", reply_markup=cancel_admin_inline_kb())

@router.message(EditStaffKPIState.kpi)
async def process_edit_staff_kpi(message: types.Message, state: FSMContext, container: Container):
    try:
        new_kpi = int(message.text)
    except ValueError:
        await message.answer("Iltimos, butun son kiriting.")
        return
        
    data = await state.get_data()
    tg_id = data.get("target_tg_id")
    user_service: UserService = container.get("user_service")
    
    updated_user = await user_service.update_user_kpi(tg_id, new_kpi)
    if updated_user:
        await message.answer(f"✅ KPI muvaffaqiyatli {new_kpi} gacha yangilandi!", reply_markup=main_admin_kb())
    else:
        await message.answer("❌ KPI yangilashda xatolik.", reply_markup=main_admin_kb())
    await state.clear()

@router.callback_query(F.data.startswith("staff_revoke_"))
async def cb_staff_revoke(call: types.CallbackQuery, container: Container):
    tg_id = int(call.data.split("_")[2])
    user_service: UserService = container.get("user_service")
    
    updated_user = await user_service.update_user_role(tg_id, UserRole.BANNED)
    if updated_user:
        await call.message.edit_text(f"🗑 Xodim faol ro'yxatdan o'chirildi. (Uni qaytarish uchun u /start bosishi kerak)")
        try:
            await call.bot.send_message(tg_id, "Profilingiz administrator tomonidan o'chirildi. Ruxsatingiz yopilgan.")
        except:
            pass
    else:
        await call.answer("Xatolik", show_alert=True)

@router.callback_query(F.data.startswith("staff_excel_"))
async def cb_staff_export_excel(call: types.CallbackQuery, container: Container):
    parts = call.data.split("_")
    period = parts[2]
    tg_id = int(parts[3])
    
    transaction_service: TransactionService = container.get("transaction_service")
    user_service: UserService = container.get("user_service")
    
    user = await user_service.get_user_by_tg_id(tg_id)
    user_pk_id = user.id if user else None
    if not user_pk_id:
        await call.answer("Foydalanuvchi topilmadi.", show_alert=True)
        return
        
    transactions = await transaction_service.get_admin_statistics(period, user_id=user_pk_id)
    if not transactions:
        await call.answer("Bu davrda tranzaksiyalar yo'q", show_alert=True)
        return
        
    import openpyxl
    from aiogram.types import BufferedInputFile
    import io
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sotuvlar"
    
    headers = ["Chek ID", "Sana/Vaqt (UTC)", "Kategoriya", "Mahsulot", "Miqdor", "Summa (so'm)"]
    ws.append(headers)
    
    for t in transactions:
        prod_name = t.product.name if t.product else "O'chirilgan"
        cat_name = t.product.category.name if getattr(t.product, 'category', None) else "Ko'rsatilmagan"
        
        row = [
            t.id,
            t.timestamp.strftime("%Y-%m-%d %H:%M:%S") if t.timestamp else "",
            cat_name,
            prod_name,
            t.amount,
            t.total_price
        ]
        ws.append(row)
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    file_name = f"Report_Staff_{user.username or tg_id}_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    file = BufferedInputFile(buf.read(), filename=file_name)
    
    await call.message.answer_document(document=file, caption=f"📥 @{user.username or tg_id} xodimi bo'yicha hisobot tayyor!")
    await call.answer()


# --- Statistics ---
@router.message(F.text == "📊 Statistika")
async def show_stats_menu(message: types.Message):
    from app.telegram.keyboards.admin import stats_periods_kb
    await message.answer("Hisobot davrini tanlang:", reply_markup=stats_periods_kb())

@router.callback_query(F.data.startswith("stats_period_"))
async def process_stats(call: types.CallbackQuery, container: Container):
    period = call.data.split("_")[2] # "today" or "week"
    transaction_service: TransactionService = container.get("transaction_service")
    
    transactions = await transaction_service.get_admin_statistics(period)
    
    if not transactions:
        await call.answer("Bu davrda sotuvlar bo'lmagan.", show_alert=True)
        return
        
    total_sales = len(transactions)
    total_items = sum(t.amount for t in transactions)
    total_revenue = sum(t.total_price for t in transactions)
    
    # Simple top 3 calculation
    from collections import defaultdict
    product_sales = defaultdict(lambda: {"count": 0, "revenue": 0})
    
    for t in transactions:
        p_name = t.product.name if t.product else f"Mahsulot {t.product_id} (o'chirilgan)"
        product_sales[p_name]["count"] += t.amount
        product_sales[p_name]["revenue"] += t.total_price
        
    top_products = sorted(product_sales.items(), key=lambda x: x[1]['count'], reverse=True)[:3]
    
    # Staff rankings
    rankings = await transaction_service.get_staff_rankings(period)
    
    period_str = "bugun" if period == "today" else "so'nggi 7 kun"
    
    text = f"📊 <b>{period_str.capitalize()} statistikasi:</b>\n\n"
    text += f"Jami cheklar: <b>{total_sales}</b>\n"
    text += f"Sotilgan mahsulotlar: <b>{total_items}</b> dona.\n"
    text += f"Umumiy daromad: <b>{total_revenue:,} so'm.</b>\n\n"
    
    text += "🏆 <b>Eng ko'p sotilgan Top-3 mahsulot:</b>\n"
    text += "<blockquote>"
    for rank, (p_name, stats) in enumerate(top_products, 1):
        text += f"{rank}. <b>{p_name}</b> — {stats['count']} dona. (<i>{stats['revenue']:,} so'm</i>)\n"
    text += "</blockquote>"
        
    if rankings:
        text += "\n👥 <b>Xodimlar reytingi:</b>\n"
        text += "<blockquote>"
        for i, rank in enumerate(rankings, 1):
            name = rank.username or f"ID {rank.tg_id}"
            text += f"{i}. <b>@{name}</b> — {rank.revenue:,} so'm. (<i>{rank.items} dona.</i>)\n"
        text += "</blockquote>"
            
    await call.message.edit_text(text, parse_mode="HTML")

# --- Rollbacks F3 ---
@router.callback_query(F.data.startswith("undo_tx_"))
async def cb_undo_tx(call: types.CallbackQuery, container: Container):
    tx_id = int(call.data.split("_")[2])
    transaction_service: TransactionService = container.get("transaction_service")
    
    success = await transaction_service.rollback_transaction(tx_id)
    if success:
        # Edit the original alert message
        import html
        original_text = html.escape(call.message.text or call.message.caption or "Tafsilotlar noma'lum")
        new_text = f"❌ <b>Tranzaksiya bekor qilindi. Mahsulot omborga qaytarildi.</b>\n\n<s>{original_text}</s>"
        await call.message.edit_text(new_text, parse_mode="HTML")
        await call.answer("Qaytarish rasmiylashtirildi!", show_alert=True)

# --- Print Receipt (Повторная печать чека) ---
@router.callback_query(F.data.startswith("print_receipt_"))
async def cb_print_receipt(call: types.CallbackQuery, container: Container):
    """Обработчик кнопки печати/повторной печати чека."""
    order_id = call.data.split("print_receipt_")[1]
    
    from app.api.printer_manager import PrinterConnectionManager
    printer_manager: PrinterConnectionManager = container.get("printer_manager")
    
    if not printer_manager.has_connected_printer:
        await call.answer(
            "⚠️ Printer hozirda ulanmagan! Printerni yoqing va qayta urinib ko'ring.",
            show_alert=True
        )
        return
    
    # Попытка повторной печати из очереди
    success = await printer_manager.retry_print_job(order_id)
    
    if success:
        await call.answer("✅ Chek printerga yuborildi!", show_alert=True)
        # Обновляем текст сообщения — убираем предупреждение
        original_text = call.message.text or ""
        if "Printer ulanmagan" in original_text:
            new_text = original_text.replace(
                "⚠️ _Printer ulanmagan. Chekni qo'lda chop etish mumkin._",
                "🖨️ _Chek printerga yuborildi._"
            )
            try:
                # Сохраняем только кнопку отмены
                from app.telegram.keyboards.admin import undo_tx_kb
                # Извлечь tx_id из кнопок (первая кнопка)
                kb = call.message.reply_markup
                await call.message.edit_text(new_text, parse_mode="Markdown", reply_markup=kb)
            except Exception:
                pass
    else:
        await call.answer(
            "❌ Chekni topib bo'lmadi yoki printerga yuborishda xatolik.",
            show_alert=True
        )


# --- Excel Export F4 ---
@router.callback_query(F.data.startswith("export_excel_"))
async def cb_export_excel(call: types.CallbackQuery, container: Container):
    period = call.data.split("_")[2]
    transaction_service: TransactionService = container.get("transaction_service")
    
    transactions = await transaction_service.get_admin_statistics(period)
    if not transactions:
        await call.answer("Bu davrda tranzaksiyalar yo'q", show_alert=True)
        return
        
    import openpyxl
    from aiogram.types import BufferedInputFile
    import io
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sotuvlar"
    
    headers = ["Chek ID", "Sana/Vaqt (UTC)", "Xodim", "Kategoriya", "Mahsulot", "Miqdor", "Summa (so'm)"]
    ws.append(headers)
    
    for t in transactions:
        prod_name = t.product.name if t.product else "O'chirilgan"
        cat_name = t.product.category.name if getattr(t.product, 'category', None) else "Ko'rsatilmagan"
        user_name = t.user.username if t.user and t.user.username else str(t.user_id)
        
        row = [
            t.id,
            t.timestamp.strftime("%Y-%m-%d %H:%M:%S") if t.timestamp else "",
            user_name,
            cat_name,
            prod_name,
            t.amount,
            t.total_price
        ]
        ws.append(row)
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    file_name = f"Report_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    file = BufferedInputFile(buf.read(), filename=file_name)
    
    await call.message.answer_document(document=file, caption="📥 Sizning Excel hisobotingiz tayyor!")
    await call.answer()

@router.callback_query(F.data == "export_inventory_excel")
async def cb_export_inventory_excel(call: types.CallbackQuery, container: Container):
    product_service: ProductService = container.get("product_service")
    products = await product_service.get_all_products()
    
    if not products:
        await call.answer("Omborda mahsulotlar yo'q.", show_alert=True)
        return
        
    import openpyxl
    from aiogram.types import BufferedInputFile
    import io
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ombor"
    
    headers = ["ID", "Kategoriya", "Nomi", "Narx", "Qoldiq", "Shtrix-kod"]
    ws.append(headers)
    
    for p in products:
        cat_name = p.category.name if p.category else "Kategoriyasiz"
        ws.append([p.id, cat_name, p.name, p.price, p.quantity, p.barcode or ""])
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    file_name = f"Inventory_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    file = BufferedInputFile(buf.read(), filename=file_name)
    
    await call.message.answer_document(document=file, caption="📦 Ombordagi qoldiqlar bo'yicha joriy hisobot tayyor!")
    await call.answer()

# --- Inventory Management (Receipt & Write-off) ---
from app.telegram.states.admin import ReceiptState, WriteOffState, BindBarcodeState
from app.telegram.keyboards.worker import worker_categories_kb, sell_product_list_kb

# Receipt Flow
@router.message(F.text == "📥 Qabul qilish")
async def start_receipt(message: types.Message, state: FSMContext, container: Container):
    category_service: CategoryService = container.get("category_service")
    categories = await category_service.get_all_categories()
    if not categories:
        await message.answer("Avvalo kategoriyalarni yarating!")
        return
    await state.set_state(ReceiptState.category_id)
    await message.answer("Qabul qilish uchun kategoriyani tanlang:", reply_markup=worker_categories_kb(categories))

@router.callback_query(ReceiptState.category_id, F.data.startswith("w_cat_"))
async def receipt_select_cat(call: types.CallbackQuery, state: FSMContext, container: Container):
    cat_id = int(call.data.split("_")[2])
    await state.update_data(category_id=cat_id)
    product_service: ProductService = container.get("product_service")
    products = await product_service.get_products_by_category(cat_id)
    if not products:
        await call.answer("Bu kategoriyada mahsulotlar yo'q.", show_alert=True)
        return
    await state.set_state(ReceiptState.product_id)
    await call.message.edit_text("Qabul qilish uchun mahsulotni tanlang:", reply_markup=sell_product_list_kb(products))

@router.callback_query(ReceiptState.product_id, F.data.startswith("sell_"))
async def receipt_select_product(call: types.CallbackQuery, state: FSMContext, container: Container):
    product_id = int(call.data.split("_")[1])
    product_service: ProductService = container.get("product_service")
    product = await product_service.get_product_by_id(product_id)
    await state.update_data(product_id=product_id, product_name=product.name)
    await call.message.edit_text(f"📦 Qabul qilish: <b>{product.name}</b>\nJoriy qoldiq: {product.quantity} dona.\n\nKirim miqdorini kiriting:", parse_mode="HTML", reply_markup=cancel_admin_inline_kb())
    await state.set_state(ReceiptState.quantity)

@router.message(ReceiptState.quantity)
async def process_receipt_quantity(message: types.Message, state: FSMContext, container: Container, db_user: User):
    try:
        qty = int(message.text)
        if qty <= 0: raise ValueError()
    except ValueError:
        await message.answer("Musbat son kiriting.")
        return
    data = await state.get_data()
    transaction_service: TransactionService = container.get("transaction_service")
    await transaction_service.create_receipt(user_id=db_user.id, product_id=data['product_id'], amount=qty)
    await message.answer(f"✅ <b>{data['product_name']}</b> mahsulotidan {qty} dona muvaffaqiyatli qabul qilindi.", parse_mode="HTML", reply_markup=main_admin_kb())
    await state.clear()

# Write-off Flow
@router.message(F.text == "🗑 Hisobdan chiqarish")
async def start_write_off(message: types.Message, state: FSMContext, container: Container):
    category_service: CategoryService = container.get("category_service")
    categories = await category_service.get_all_categories()
    if not categories:
        await message.answer("Avvalo kategoriyalarni yarating!")
        return
    await state.set_state(WriteOffState.category_id)
    await message.answer("Hisobdan chiqarish uchun kategoriyani tanlang:", reply_markup=worker_categories_kb(categories))

@router.callback_query(WriteOffState.category_id, F.data.startswith("w_cat_"))
async def write_off_select_cat(call: types.CallbackQuery, state: FSMContext, container: Container):
    cat_id = int(call.data.split("_")[2])
    await state.update_data(category_id=cat_id)
    product_service: ProductService = container.get("product_service")
    products = await product_service.get_products_by_category(cat_id)
    available = [p for p in products if p.quantity > 0]
    if not available:
        await call.answer("Hisobdan chiqarish uchun mavjud mahsulotlar yo'q.", show_alert=True)
        return
    await state.set_state(WriteOffState.product_id)
    await call.message.edit_text("Hisobdan chiqarish uchun mahsulotni tanlang:", reply_markup=sell_product_list_kb(available))

@router.callback_query(WriteOffState.product_id, F.data.startswith("sell_"))
async def write_off_select_product(call: types.CallbackQuery, state: FSMContext, container: Container):
    product_id = int(call.data.split("_")[1])
    product_service: ProductService = container.get("product_service")
    product = await product_service.get_product_by_id(product_id)
    await state.update_data(product_id=product_id, product_name=product.name, max_qty=product.quantity)
    await call.message.edit_text(f"🗑 Hisobdan chiqarish: <b>{product.name}</b>\nMavjud: {product.quantity} dona.\n\nHisobdan chiqarish miqdorini kiriting:", parse_mode="HTML", reply_markup=cancel_admin_inline_kb())
    await state.set_state(WriteOffState.quantity)

@router.message(WriteOffState.quantity)
async def process_write_off_quantity(message: types.Message, state: FSMContext):
    try:
        qty = int(message.text)
        if qty <= 0: raise ValueError()
    except ValueError:
        await message.answer("Musbat son kiriting.")
        return
    data = await state.get_data()
    if qty > data['max_qty']:
        await message.answer(f"Ombordagidan ({data['max_qty']}) ko'prog'ini hisobdan chiqarib bo'lmaydi. Qaytadan kiriting:")
        return
    await state.update_data(quantity=qty)
    await message.answer("Hisobdan chiqarish sababini kiriting (masalan, 'yaroqsiz', 'muddati o'tgan'):")
    await state.set_state(WriteOffState.reason)

@router.message(WriteOffState.reason)
async def process_write_off_reason(message: types.Message, state: FSMContext, container: Container, db_user: User):
    reason = message.text.strip()
    data = await state.get_data()
    transaction_service: TransactionService = container.get("transaction_service")
    await transaction_service.create_write_off(user_id=db_user.id, product_id=data['product_id'], amount=data['quantity'], reason=reason)
    await message.answer(f"✅ <b>{data['product_name']}</b> mahsulotidan {data['quantity']} dona muvaffaqiyatli hisobdan chiqarildi, sabab: {reason}", parse_mode="HTML", reply_markup=main_admin_kb())
    await state.clear()

# --- Bind Barcode ---
@router.callback_query(F.data.startswith("prod_barcode_"))
async def cb_bind_barcode(call: types.CallbackQuery, state: FSMContext):
    product_id = int(call.data.split("_")[2])
    await state.update_data(product_id=product_id)
    await state.set_state(BindBarcodeState.barcode)
    await call.message.edit_text("Mahsulot shtrix-kodini yuboring (raqamlarni yozing yoki kamera yordamida skanerlang):", reply_markup=cancel_admin_inline_kb())

@router.message(BindBarcodeState.barcode)
async def process_bind_barcode(message: types.Message, state: FSMContext, container: Container):
    barcode = message.text.strip()
    data = await state.get_data()
    product_service: ProductService = container.get("product_service")
    
    existing = await product_service.get_product_by_barcode(barcode)
    if existing and existing.id != data['product_id']:
        await message.answer("Bu shtrix-kod boshqa mahsulotga biriktirilgan. Boshqasini sinab ko'ring:")
        return
        
    success = await product_service.update_barcode(data['product_id'], barcode)
    if success:
        await message.answer(f"✅ Shtrix-kod {barcode} muvaffaqiyatli biriktirildi!", reply_markup=main_admin_kb())
    else:
        await message.answer("Shtrix-kod biriktirishda xatolik.", reply_markup=main_admin_kb())
    await state.clear()

